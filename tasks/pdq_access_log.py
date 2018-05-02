"""
Report on which data partners fetched PDQ data.
"""

# Standard library modules
import argparse
import datetime
import gzip
import os
import re

# Third-party modules
from dateutil.relativedelta import relativedelta
from lxml import etree
import openpyxl
import requests
from core.exceptions import TaskException
from task_property_bag import TaskPropertyBag

# Project modules
import cdr
from cdr_task_base import CDRTask

class Report(CDRTask):
    """
    Task for generating a spreadsheet showing which PDQ partners
    have connected to the SFTP server to retrieve data.

    Attributes:
        logger - object for recording what we do
        resend - if True send a previously saved report
        noemail - if True don't email the report
        recips - sequence of recipient email addresses
        month - period for which activity is reported
        log_path - location of the log which we parse
        report_path - location to which report file is written
    """

    LOGNAME = "pdq-access-report"
    SENDER = "NCI PDQ Operator <NCIPDQoperator@mail.nih.gov>"
    SUBJECT = "SFTP Log - PDQ Distribution Partner Access Report (%s)"
    MAX_TRIES = 5
    DELAY = 5
    WIDTHS = 15, 50, 40, 10, 10, 10
    LABELS = "Login", "Partner", "Path", "Session", "Date", "Time"
    NON_PARTNERS = set(cdr.getControlValue("PDQ",
                                           "non-partners", "").split(","))

    def __init__(self, parms, data):
        """
        Create a logger and set up the options.

        level
            optional logging level; defaults to "info"
        resend
            if True just send an already generated report
        noemail
            if True do everything except email the report (for testing)
        month
            override YYYYMM (default is previous month)
        recips
            comma-separated email addresses (default is content
            dissemination mailbox)
        """

        CDRTask.__init__(self, parms, data)
        self.logger.info("Report started")
        self.set_opts(parms)

    def set_opts(self, parms):
        """
        Extract the runtime options for this task.
        """

        for name in parms:
            self.logger.info("Option %s=%r", name, parms[name])
        self.resend = parms.get("resend") and True or False
        self.noemail = parms.get("noemail") and True or False
        self.month = self.Month(parms.get("month"))
        self.recips = self.get_recips(parms.get("recips"))
        self.report_path = self.month.report_path()
        self.log_path = self.month.log_path()

    def get_recips(self, recips):
        """
        Figure out who we should send the report to.
        """

        if recips:
            return [recip.strip() for recip in recips.split(",")]
        return ["NCIContentDissemination@nih.gov"]

    def Perform(self):
        """
        Generate and/or send the report.
        """

        if not self.resend:
            self.orgs = self.get_orgs()
            requests = self.get_requests()
            self.make_report(requests)
        if not self.noemail:
            self.send_report()
        return TaskPropertyBag()

    def get_orgs(self):
        """
        Fetch the information about the organizations with which we partner.
        """

        url = "https://cdr.cancer.gov/cgi-bin/cdr/get-pdq-partners.py?p=CDR"
        self.logger.info("fetching partners from %r", url)
        class Org:
            def __init__(self, node):
                self.oid = int(node.get("oid"))
                self.name = cdr.get_text(node.find("org_name"))
                self.status = cdr.get_text(node.find("org_status"))
                self.uid = cdr.get_text(node.find("ftp_userid"))
                self.terminated = cdr.get_text(node.find("terminated"))
        root = etree.fromstring(requests.get(url).content)
        orgs = {}
        for node in root.findall("org_id"):
            org = Org(node)
            if org.uid is not None:
                orgs[org.uid] = org
        return orgs

    def get_requests(self):
        """
        Parse the log file, extracting the partner requests.

        Make sure we have the latest log files (using rsync),
        and then walk through each line in the log file for this
        report. We're interested in two types of lines (records):
        session opening lines, from which we build our dictionary
        of login IDs mapped by session IDs; and file opening lines,
        from which we parse our request objects. We skip over requests
        made using login accounts which are known not to represent
        PDQ data partners (CBIIT accounts, developer accounts, testing
        accounts, etc.).

        The session IDs appear in fields which look like this example:

            sshd[9223]:

        ... which is why we use the expression [5:-2] to extract them.
        """

        class Request:
            def __init__(self, line, sids, orgs):
                """
                Extract the fields from the sftp activity log.

                Note that the second field holds the digit(s) for
                the date the request was received. In order to
                ensure that the value has a uniform width (for
                possible sorting purposes), we stick a zero in
                front of the value and use the substring starting
                two characters from the end (hence the -2 in the
                tokens[1][-2:] expression).

                Passed:
                  line - record from the sftp log, fields separated by spaces
                  sids - dictionary of sftp login IDs indexed by session ID
                  orgs - dictionary of partner org names indexed by login ID
                """

                tokens = line.split()
                self.date = "%s-%s" % (tokens[0], ("0" + tokens[1])[-2:])
                self.time = tokens[2]
                self.path = tokens[6][1:-1].replace("/pdq/full/", "")
                self.sid = int(tokens[4][5:-2])
                self.user = sids.get(self.sid, "")
                if self.user and self.user in orgs:
                    self.org = orgs[self.user].name or ""
                else:
                    self.org = ""
        requests = {}
        sids = {}
        count = 0
        self.logger.info("parsing %r", self.log_path)
        self.sync_logs()
        with gzip.open(self.log_path) as fp:
            for line in fp.readlines():
                if "]: open " in line:
                    request = Request(line, sids, self.orgs)
                    if request.user in self.NON_PARTNERS:
                        continue
                    if request.user not in requests:
                        requests[request.user] = []
                    requests[request.user].append(request)
                    count += 1
                elif "session opened for local user" in line:
                    tokens = line.split()
                    sid = int(tokens[4][5:-2])
                    user = tokens[10]
                    sids[sid] = user
        args = count, len(requests)
        self.logger.info("fetched %d requests from %d partners", *args)
        return requests

    def sync_logs(self):
        """
        Top up our local copies of the pdq logs from the sftp server.
        We're ignoring some expected errors, having to do with cygwin's
        difficulty in dealing with bizarre Windows file permissions
        configuration settings. If we really fail to bring down a needed
        log file successfully, we'll find out when we try to read it.
        """

        ssh = ("ssh -i d:/etc/cdroperator_rsa -o LogLevel=error "
               "-o StrictHostKeyChecking=no")
        usr = "cdroperator"
        dns = "cancerinfo.nci.nih.gov"
        src = "%s@%s:/sftp/sftphome/cdrstaging/logs/*" % (usr, dns)
        cmd = "rsync -e \"%s\" %s ." % (ssh, src)
        fix = r"%s:\cdr\bin\fix-permissions.cmd ." % cdr.WORK_DRIVE
        cwd = os.getcwd()
        os.chdir(self.Month.LOGDIR)
        self.logger.info(cmd)
        cdr.runCommand(cmd)
        self.logger.info(fix)
        cdr.runCommand(fix)
        os.chdir(cwd)

    def make_report(self, requests):
        """
        Generate and save a report of files fetched by the PDQ partners.
        """

        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Requests"
        bold = openpyxl.styles.Font(size=12, bold=True)
        center = openpyxl.styles.Alignment(horizontal="center")
        sheet.freeze_panes = "A6"
        sheet["A1"] = str(datetime.date.today())
        sheet["A1"].font = bold
        sheet["A3"] = "Downloads for %s" % self.month
        sheet["A3"].font = bold
        for i, width in enumerate(self.WIDTHS):
            col = chr(ord("A")+i)
            cell = "%s5" % col
            sheet.column_dimensions[col].width = width
            sheet[cell] = self.LABELS[i]
            sheet[cell].font = bold
            sheet[cell].alignment = center
        row = 6
        for user in sorted(requests):
            for r in requests[user]:
                sheet.cell(row=row, column=1, value=r.user)
                sheet.cell(row=row, column=2, value=r.org)
                sheet.cell(row=row, column=3, value=r.path)
                sheet.cell(row=row, column=4, value=r.sid)
                sheet.cell(row=row, column=5, value=r.date).alignment=center
                sheet.cell(row=row, column=6, value=r.time).alignment=center
                row += 1
        book.save(self.report_path)
        self.logger.info("wrote %r", self.report_path)

    def send_report(self):
        """
        Send the report as an attachment to an email message.
        """

        label = str(self.month)
        book = cdr.EmailAttachment(filepath=self.report_path)
        subject = self.SUBJECT % label
        body = (
            "Attached is the monthly PDQ Distribution Partner report listing "
            "all documents downloaded from the SFTP server for %s.\n" % label,
            "The report is based on the log file provided at",
            "         %s\n" % self.log_path,
            "Please save the attached report to the network directory",
            "         L:\\OCPL\\_CROSS\\CDR\\Reports\\FTP Stats",
            "so the Clinical Trials team can access the information as needed.",
            "",
            "For questions or comments please reply to this email message."
        )
        body = "\n".join(body)
        recips = self.recips
        cdr.sendMail(self.SENDER, recips, subject, body, attachments=[book])
        self.logger.info("sent report to %s", ", ".join(recips))

    class Month:
        """
        Period covered by the report.

        Attributes:
            year - integer for the year of the report's data
            month - integer for the month of the report's data
        """

        LOGDIR = cdr.BASEDIR + "/sftp_log"
        REPORTS = cdr.BASEDIR + "/reports"
        FILEBASE = "PDQPartnerDownloads"

        def __init__(self, yyyymm=None):
            """
            Extract the month and year from the YYYYMM string passed,
            if provided; otherwise get the month and year for the
            previous month.
            """

            if yyyymm:
                match = re.match(r"(\d\d\d\d)(\d\d)", yyyymm)
                if not match:
                    self.logger.error("malformed month %r", month)
                    raise Exception("invalid parameter")
                self.year = int(match.group(1))
                self.month = int(match.group(2))
                self.start = datetime.date(self.year, self.month, 1)
            else:
                today = datetime.date.today()
                self.start = today - relativedelta(months=1, day=1)
                self.month = self.start.month
                self.year = self.start.year

        def log_path(self):
            """
            Contruct the path for the location of the log to be parsed.
            """

            report_date = self.start + relativedelta(months=1)
            stamp = report_date.strftime("%Y%m%d")
            return "%s/pdq.log-%s.gz" % (self.LOGDIR, stamp)

        def report_path(self):
            """
            Construct the path for the location of the report to be generated.
            """

            ym = self.start.strftime("%Y-%m")
            return "%s/%s_%s.xlsx" % (self.REPORTS, self.FILEBASE, ym)

        def __str__(self):
            """
            Display the month in spelled-out English.

            This is invoked by the make_report() method to construct
            the report title. Python effectively calls str(arg) when
            interpolating arguments for '%s' placeholders in the string
            templates, and the built-in str() function uses the object's
            __str__() method if it has one.
            """

            return self.start.strftime("%B %Y")

if __name__ == "__main__":
    """
    Support command-line testing.
    """

    parser = argparse.ArgumentParser()
    parser.add_argument("--level", default="debug")
    parser.add_argument("--noemail", action="store_true")
    parser.add_argument("--resend", action="store_true")
    parser.add_argument("--month")
    parser.add_argument("--recips")
    Report(vars(parser.parse_args()), {}).Perform()
